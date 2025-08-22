from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from decimal import Decimal

class ExcelReportGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def apply_header_style(self, worksheet, row_num, col_start, col_end):
        """Appliquer le style d'en-tête"""
        for col in range(col_start, col_end + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
    
    def apply_data_style(self, worksheet, row_start, row_end, col_start, col_end):
        """Appliquer le style aux données"""
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = self.center_alignment
    
    def auto_adjust_columns(self, worksheet):
        """Ajuster automatiquement la largeur des colonnes"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def generate_daily_report_excel(self, daily_report, sales_data, stock_alerts):
        """Générer un rapport quotidien en Excel"""
        wb = Workbook()
        
        # Feuille 1: Résumé
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        # Titre
        ws_summary['A1'] = f"Rapport Quotidien - {daily_report.date.strftime('%d/%m/%Y')}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:D1')
        
        # Date de génération
        ws_summary['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws_summary.merge_cells('A2:D2')
        
        # Résumé des ventes
        ws_summary['A4'] = "RÉSUMÉ DES VENTES"
        ws_summary['A4'].font = Font(bold=True, size=12)
        
        summary_data = [
            ['Indicateur', 'Valeur'],
            ['Nombre de ventes', daily_report.total_sales_count],
            ['Chiffre d\'affaires (BIF)', f"{daily_report.total_sales:,.0f}"],
            ['Bénéfice brut (BIF)', f"{daily_report.total_profit:,.0f}"],
            ['Ticket moyen (BIF)', f"{daily_report.average_sale:,.0f}"],
            ['Dépenses (BIF)', f"{daily_report.total_expenses:,.0f}"],
            ['Résultat net (BIF)', f"{daily_report.net_result:,.0f}"]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        
        self.apply_header_style(ws_summary, 5, 1, 2)
        self.apply_data_style(ws_summary, 6, 11, 1, 2)
        
        # Feuille 2: Détail des ventes
        if sales_data:
            ws_sales = wb.create_sheet(title="Ventes Détaillées")
            
            # En-têtes
            headers = ['Produit', 'Catégorie', 'Quantité', 'Prix Unitaire (BIF)', 'Montant Total (BIF)']
            for col_idx, header in enumerate(headers, start=1):
                ws_sales.cell(row=1, column=col_idx, value=header)
            
            # Données
            for row_idx, sale in enumerate(sales_data, start=2):
                ws_sales.cell(row=row_idx, column=1, value=sale['product_name'])
                ws_sales.cell(row=row_idx, column=2, value=sale.get('category_name', ''))
                ws_sales.cell(row=row_idx, column=3, value=sale['quantity'])
                ws_sales.cell(row=row_idx, column=4, value=float(sale['unit_price']))
                ws_sales.cell(row=row_idx, column=5, value=float(sale['total_amount']))
            
            self.apply_header_style(ws_sales, 1, 1, 5)
            self.apply_data_style(ws_sales, 2, len(sales_data) + 1, 1, 5)
            self.auto_adjust_columns(ws_sales)
        
        # Feuille 3: Alertes de stock
        if stock_alerts:
            ws_alerts = wb.create_sheet(title="Alertes Stock")
            
            # En-têtes
            headers = ['Produit', 'Catégorie', 'Stock Actuel', 'Stock Minimum', 'Type d\'Alerte', 'Statut']
            for col_idx, header in enumerate(headers, start=1):
                ws_alerts.cell(row=1, column=col_idx, value=header)
            
            # Données
            for row_idx, alert in enumerate(stock_alerts, start=2):
                ws_alerts.cell(row=row_idx, column=1, value=alert.product.name)
                ws_alerts.cell(row=row_idx, column=2, value=alert.product.category.name)
                ws_alerts.cell(row=row_idx, column=3, value=alert.product.current_stock)
                ws_alerts.cell(row=row_idx, column=4, value=alert.product.minimum_stock)
                ws_alerts.cell(row=row_idx, column=5, value=alert.get_alert_type_display())
                ws_alerts.cell(row=row_idx, column=6, value=alert.get_status_display())
            
            self.apply_header_style(ws_alerts, 1, 1, 6)
            self.apply_data_style(ws_alerts, 2, len(stock_alerts) + 1, 1, 6)
            self.auto_adjust_columns(ws_alerts)
        
        self.auto_adjust_columns(ws_summary)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_stock_report_excel(self, products):
        """Générer un rapport de stock en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport de Stock"
        
        # Titre
        ws['A1'] = f"Rapport de Stock - {timezone.now().strftime('%d/%m/%Y')}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        
        # En-têtes
        headers = ['Produit', 'Catégorie', 'Stock Actuel', 'Stock Minimum', 'Prix d\'Achat (BIF)', 'Valeur Stock (BIF)', 'Statut']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header)
        
        # Données
        total_value = Decimal('0.00')
        low_stock_count = 0
        
        for row_idx, product in enumerate(products, start=4):
            stock_value = product.current_stock * product.purchase_price
            total_value += stock_value
            
            status = "⚠️ Stock Faible" if product.current_stock <= product.minimum_stock else "✅ OK"
            if product.current_stock <= product.minimum_stock:
                low_stock_count += 1
            
            ws.cell(row=row_idx, column=1, value=product.name)
            ws.cell(row=row_idx, column=2, value=product.category.name)
            ws.cell(row=row_idx, column=3, value=product.current_stock)
            ws.cell(row=row_idx, column=4, value=product.minimum_stock)
            ws.cell(row=row_idx, column=5, value=float(product.purchase_price))
            ws.cell(row=row_idx, column=6, value=float(stock_value))
            ws.cell(row=row_idx, column=7, value=status)
        
        # Statistiques en bas
        stats_row = len(products) + 6
        ws.cell(row=stats_row, column=1, value="STATISTIQUES GÉNÉRALES").font = Font(bold=True)
        ws.cell(row=stats_row + 1, column=1, value="Nombre total de produits:")
        ws.cell(row=stats_row + 1, column=2, value=len(products))
        ws.cell(row=stats_row + 2, column=1, value="Produits en stock faible:")
        ws.cell(row=stats_row + 2, column=2, value=low_stock_count)
        ws.cell(row=stats_row + 3, column=1, value="Valeur totale du stock (BIF):")
        ws.cell(row=stats_row + 3, column=2, value=float(total_value))
        
        # Styles
        self.apply_header_style(ws, 3, 1, 7)
        self.apply_data_style(ws, 4, len(products) + 3, 1, 7)
        self.auto_adjust_columns(ws)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_sales_report_excel(self, sales, start_date, end_date):
        """Générer un rapport de ventes en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport de Ventes"
        
        # Titre
        period_str = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        ws['A1'] = f"Rapport de Ventes - {period_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # En-têtes
        headers = ['Date', 'Table', 'Produit', 'Quantité', 'Prix Unitaire (BIF)', 'Total (BIF)', 'Serveur', 'Statut']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header)
        
        # Données
        total_amount = Decimal('0.00')
        
        for row_idx, sale in enumerate(sales, start=4):
            for item in sale.items.all():
                ws.cell(row=row_idx, column=1, value=sale.created_at.strftime('%d/%m/%Y %H:%M'))
                ws.cell(row=row_idx, column=2, value=sale.table.number if sale.table else 'N/A')
                ws.cell(row=row_idx, column=3, value=item.product.name)
                ws.cell(row=row_idx, column=4, value=item.quantity)
                ws.cell(row=row_idx, column=5, value=float(item.unit_price))
                ws.cell(row=row_idx, column=6, value=float(item.total_price))
                ws.cell(row=row_idx, column=7, value=sale.server.get_full_name())
                ws.cell(row=row_idx, column=8, value=sale.get_status_display())
                
                total_amount += item.total_price
                row_idx += 1
        
        # Total en bas
        total_row = row_idx + 1
        ws.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(total_amount)).font = Font(bold=True)
        
        # Styles
        self.apply_header_style(ws, 3, 1, 8)
        self.apply_data_style(ws, 4, row_idx - 1, 1, 8)
        self.auto_adjust_columns(ws)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
